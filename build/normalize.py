# -*- coding: utf-8 -*-
"""
대학알리미 4년제 학과 리스트(2014-2026) 정규화 파이프라인.

산출:
  data/records.json   - 정규화/필터링된 활성 학과 레코드(연도별 스냅샷 + 상태 플래그)
  data/schools.json   - 캐노니컬 학교 식별(코드 기반 + 명칭변경/통합 lineage)
  data/summary.json   - 연도/계열/필터 통계 (검증용)

규칙(사용자 확정):
  - 학교범위: 대학교 + 교육대학 + (전환된)산업대학. 사이버/방송통신/각종학교/기술대학 제외.
  - 무전공/자유전공/자율전공/광역·계열모집 전부 제외.
  - 교양대학/교양학부/기초교육 제외. 1학년전용학과 제외.
  - 국립화 명칭변경(OO대->국립OO대, 안동대->국립경국대)은 학교코드로 동일학교 취급(이벤트X).
  - 학교 간 통합(경상국립대=경상대+경남과기대, 강원대+강릉원주대)만 merge event.
"""
import openpyxl, json, re, collections, os

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
YEARS = list(range(2014, 2027))

# ---- 학교구분(학교유형) 화이트리스트 ----
# 2026은 학교구분='대학'로 뭉뚱그려져 있고 실제 유형은 '학제' 컬럼에 있음 -> 학제 우선.
SCHOOL_TYPE_KEEP = {"대학교", "산업대학", "교육대학"}

# ---- 컬럼 별칭(연도별 헤더 변동 흡수) ----
ALIASES = {
    "year":   ["조사년도"],
    "stype":  ["학제", "학교구분"],
    "school": ["학교명"],
    "scode":  ["학교코드"],
    "college":["단과대학명", "단과대학"],
    "dept":   ["학부·과(전공)명"],
    "dcode":  ["학교별학과코드"],
    "char":   ["학과특성"],
    "status": ["학과상태"],
    "broad":  ["표준분류대계열", "표준대계열", "대계열분류", "대계열", "표준분류계열"],
    "mid":    ["표준분류중계열", "표준중계열", "중계열분류", "중계열"],
    "sub":    ["표준분류소계열", "표준소계열", "소계열분류", "소계열"],
}

def _norm_header(c):
    return str(c).replace("\n", "").strip() if c is not None else ""

def load_year(y):
    wb = openpyxl.load_workbook(os.path.join(ROOT, f"{y}.xlsx"), read_only=True)
    ws = wb[wb.sheetnames[0]]
    hr = hdr = None
    for i, r in enumerate(ws.iter_rows(min_row=1, max_row=12, values_only=True)):
        if any(c and "학교명" in str(c) for c in r):
            hr = i
            hdr = [_norm_header(c) for c in r]
            break
    idx = {h: i for i, h in enumerate(hdr)}
    def col(key):
        for a in ALIASES[key]:
            if a in idx:
                return idx[a]
        return None
    cmap = {k: col(k) for k in ALIASES}
    rows = []
    for r in ws.iter_rows(min_row=hr + 2, values_only=True):
        if not r or not any(r):
            continue
        rows.append(r)
    wb.close()
    return cmap, rows

def get(r, cmap, key):
    i = cmap.get(key)
    if i is None or i >= len(r) or r[i] is None:
        return ""
    return str(r[i]).strip()

# ---- 상태 분류 ----
GONE = ("폐지", "폐과")  # 해당 연도에 사라진(신입생 미모집) 학과
def status_class(st):
    """returns (present:bool, event:set)"""
    present = not any(g in st for g in GONE)
    ev = set()
    if "신설" in st: ev.add("new")
    if "통합" in st: ev.add("merge")
    if "분리" in st: ev.add("split")
    if "변경" in st: ev.add("change")
    return present, ev

# ---- 제외(무전공/교양/1학년전용) ----
RE_MUJEONGONG = re.compile(
    r"자유전공|자율전공|무전공|무학과|광역|계열모집|"
    r"^(인문|자연|사회|공학|예체능|예술|자연과학|인문사회|이공|융합)\s*계열$"
)
RE_GYOYANG = re.compile(r"교양|학부대학|기초교육|기초교양|기초과학부|기초의과학|글로벌리더|자유전공")
RE_FIRSTYEAR = re.compile(r"1학년|일학년|신입생|예과군|학부\(1")

def canon_broad(b):
    """대계열 라벨 표기변동(인문ㆍ사회/인문사회계열 등) 통일."""
    b = (b or "").replace(" ", "")
    if "광역" in b: return "광역"            # 무전공/광역모집 -> 제외 대상
    if "인문" in b or "사회" in b: return "인문사회"
    if "공학" in b: return "공학"
    if "자연" in b: return "자연과학"
    if "예" in b and "체" in b: return "예체능"
    if "의학" in b or "의약" in b or "약학" in b: return "의학"
    return b or "기타"

def is_excluded(dept, college):
    d = dept.replace(" ", "")
    c = (college or "").replace(" ", "")
    if RE_MUJEONGONG.search(d):
        return "무전공"
    if RE_FIRSTYEAR.search(d):
        return "1학년전용"
    # 교양: 학과명 또는 단과대학이 교양/학부대학 계열
    if RE_GYOYANG.search(d) or RE_GYOYANG.search(c):
        return "교양"
    return None

def main():
    # 1) 코드->연도별 명칭 (2021+) 으로 캐노니컬 학교명 도출
    raw = {y: load_year(y) for y in YEARS}
    code_names = collections.defaultdict(dict)   # code -> {year: name}
    name_type = {}                                # (year,name) -> stype
    for y in YEARS:
        cmap, rows = raw[y]
        for r in rows:
            sc = get(r, cmap, "scode")
            nm = get(r, cmap, "school")
            st = get(r, cmap, "stype")
            name_type[(y, nm)] = st
            if sc:
                code_names[sc][y] = nm
    # 캐노니컬명 = 코드의 가장 최근 연도 명칭 (코드 있는 경우)
    code_canon = {}
    name_to_code = {}   # (year,name)->code  (있으면)
    for sc, yn in code_names.items():
        latest = yn[max(yn)]
        code_canon[sc] = latest
        for y, nm in yn.items():
            name_to_code[(y, nm)] = sc

    # 코드 없는 연도(2014-2020) 명칭을 캐노니컬로 잇기:
    # 동일 명칭이 2021+에 존재하면 그 코드의 캐노니컬명 사용.
    name2021plus = {}  # name -> code (2021+ 중 가장 이른)
    for sc, yn in code_names.items():
        for y in sorted(yn):
            name2021plus.setdefault(yn[y], sc)

    # 수동 lineage: 국립화/통합 이전 옛 명칭 -> 코드(또는 캐노니컬)
    # (국립화 명칭변경: 동일학교 / 통합: merge event 별도)
    MANUAL_RENAME = {
        "경상대학교": "경상국립대학교",         # 국립화+통합 주체(메인) = 동일학교
        "서울과학기술대학교(산업대)": "서울과학기술대학교",
        # 분교 -> 존속 캠퍼스 개명(같은 캠퍼스 식별, 통합 아님)
        "연세대학교(원주)": "연세대학교(미래)",      # 2020 개명, 미래캠 존속
        "한국전통문화대학교(일반)": "한국전통문화대학교",
        # 2026 신설 캠퍼스(글로컬 통합)는 본교 소속으로
        "국립목포대학교(담양캠퍼스)": "국립목포대학교",
        "국립창원대학교(거창캠퍼스)": "국립창원대학교",
        "국립창원대학교(남해캠퍼스)": "국립창원대학교",
    }
    # 통합으로 흡수된 학교 -> 흡수 주체, 통합연도 (캐노니컬 명칭 기준)
    # 흡수된 학교는 흡수 주체 페이지에서 하단 밴드로 합쳐 표시(독립목록에선 제외).
    MERGERS = {
        "경남과학기술대학교": {"into": "경상국립대학교", "year": 2022},
        "국립강릉원주대학교": {"into": "강원대학교",     "year": 2026},
        # 분교가 본교 2캠퍼스로 흡수 = 사실상 학교통합
        "상명대학교(천안)":   {"into": "상명대학교",     "year": 2017},
        "홍익대학교(세종)":   {"into": "홍익대학교",     "year": 2017},
    }

    def canon_school(y, name):
        if name in MANUAL_RENAME:
            name = MANUAL_RENAME[name]
        sc = name_to_code.get((y, name)) or name2021plus.get(name)
        if sc:
            return code_canon[sc]
        return name

    # 2) 레코드 생성 + 필터
    records = []
    excl_counter = collections.Counter()
    type_drop = collections.Counter()
    for y in YEARS:
        cmap, rows = raw[y]
        for r in rows:
            stype = get(r, cmap, "stype")
            if stype not in SCHOOL_TYPE_KEEP:
                type_drop[stype] += 1
                continue
            dept = get(r, cmap, "dept")
            if not dept:
                continue
            college = get(r, cmap, "college")
            st = get(r, cmap, "status")
            present, ev = status_class(st)
            broad = canon_broad(get(r, cmap, "broad"))
            if broad == "광역":            # 무전공/광역모집 broad -> 제외
                excl_counter["무전공"] += 1
                continue
            ex = is_excluded(dept, college)
            if ex:
                excl_counter[ex] += 1
                continue
            school_raw = get(r, cmap, "school")
            rec = {
                "year": y,
                "school_raw": school_raw,
                "school": canon_school(y, school_raw),
                "code": get(r, cmap, "scode"),
                "dcode": get(r, cmap, "dcode"),
                "college": college,
                "dept": dept,
                "status": st,
                "present": present,
                "event": sorted(ev),
                "broad": broad,
                "mid": get(r, cmap, "mid"),
                "sub": get(r, cmap, "sub"),
            }
            records.append(rec)

    # 2.5) 2026는 중/소계열 미수록(대계열만) -> 직전(<=2025) 동일(학교,학과명)에서 상속
    prior = {}
    for rec in records:
        if rec["year"] <= 2025 and rec["sub"]:
            prior[(rec["school"], rec["dept"])] = (rec["sub"], rec["mid"])
    inherited = 0
    for rec in records:
        if rec["year"] == 2026:
            pm = prior.get((rec["school"], rec["dept"]))
            if pm:
                rec["sub"], rec["mid"] = pm; inherited += 1
            else:
                rec["sub"] = ""        # 신규/매칭없음: 소계열 미상 -> broad 폴백

    # 3) merge candidate auto-detect: 학교(캐노니컬)별 활성 학과수가 직전연도>0 -> 0
    active_by_school = collections.defaultdict(lambda: collections.defaultdict(int))
    for rec in records:
        if rec["present"]:
            active_by_school[rec["school"]][rec["year"]] += 1
    merge_candidates = []
    for sch, ys in active_by_school.items():
        for y in YEARS[1:]:
            prev = ys.get(y - 1, 0)
            cur = ys.get(y, 0)
            if prev >= 10 and cur == 0:
                merge_candidates.append({"school": sch, "last_active_year": y - 1, "dropped_year": y})

    summary = {
        "years": YEARS,
        "n_records_active": sum(1 for r in records if r["present"]),
        "n_records_total": len(records),
        "excluded": dict(excl_counter),
        "school_type_dropped": dict(type_drop),
        "merge_candidates(auto)": merge_candidates,
        "mergers(manual)": MERGERS,
        "active_dept_count_by_year": {
            y: sum(1 for r in records if r["present"] and r["year"] == y) for y in YEARS
        },
        "n_schools_by_year": {
            y: len({r["school"] for r in records if r["present"] and r["year"] == y}) for y in YEARS
        },
    }

    os.makedirs(os.path.join(ROOT, "data"), exist_ok=True)
    with open(os.path.join(ROOT, "data", "records.json"), "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False)
    with open(os.path.join(ROOT, "data", "schools.json"), "w", encoding="utf-8") as f:
        json.dump({"code_canon": code_canon, "mergers": MERGERS, "manual_rename": MANUAL_RENAME},
                  f, ensure_ascii=False, indent=2)
    with open(os.path.join(ROOT, "data", "summary.json"), "w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)

    print(json.dumps(summary, ensure_ascii=False, indent=2))

if __name__ == "__main__":
    main()
